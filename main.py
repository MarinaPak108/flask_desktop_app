import os
from flaskwebgui import FlaskUI
from flask import Flask, render_template
import random
import openpyxl

app = Flask(__name__, static_url_path="", static_folder=os.path.join("C:", "/phrases"))

directory = os.getcwd()

@app.route("/")
def home():
    path=os.path.join("C:", "/phrases/names.xlsx")
    files= os.listdir("/phrases")
    res = list(filter(lambda x: all(y not in "." for y in x), files))
    wb = openpyxl.load_workbook(path)
    names = wb.sheetnames
    print(names)
    return render_template('home.html',fileNames = zip(res,names))

@app.route("/greet/<file>/<name>")
def greet(file='a', name='common'):
    f_path = os.path.join('/phrases/', file)
    images =  os.listdir(f_path)
    if len(images) == 0: 
        image_path = '/img.jpeg'
    else:
        image_path = os.path.join('/'+file+"/", random.choice(images))
    path=os.path.join("C:", "/phrases/names.xlsx")
    wb = openpyxl.load_workbook(path)
    sheet = wb[name] 
    phrases = list(map(lambda v:v.value, sheet['A']))
    if phrases[0] == None:
        phrase = "not ready yet"
    else:
        phrase = random.choice(phrases)
    return render_template('greet.html',img=image_path, phrase = phrase, file = file, name = name)

if __name__ == "__main__":
  FlaskUI(app=app, server="flask",  width= 700, height=700).run()
  